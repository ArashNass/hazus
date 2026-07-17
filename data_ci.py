"""
data_ci.py
Parses Nirandjan et al. (2024) Critical Infrastructure Vulnerability Database.

Two fundamentally different curve types:
  Vulnerability sheets (F_Vuln_Depth, E_Vuln_PGA, W_Vuln_V10m, L_Vuln_PGD):
    - One column per curve ID, header = 'Damage factor'
    - Y-axis: damage ratio 0-1
    - Stored as: {x: [...], y: [...], y_lo, y_hi, curve_type='Vulnerability'}

  Fragility sheets (E_Frag_PGA, L_Frag_PGD):
    - Same ID repeated for each damage state (Slight/Moderate/Extensive/Complete)
    - Y-axis: probability of exceedance 0-1
    - Stored as: {x: [...], ds: {Slight:[...], Moderate:[...], ...}, curve_type='Fragility'}

Licence: CC BY 4.0
Cite: Nirandjan et al. (2024), NHESS, doi:10.5194/nhess-24-4341-2024
"""

import os, re
import openpyxl

CI_SYSTEM = {
    '1':'Energy','2':'Energy','3':'Energy','4':'Energy','5':'Energy','6':'Energy',
    '7':'Transport','8':'Transport','9':'Transport',
    '10':'Telecom','11':'Telecom','12':'Telecom',
    '13':'Water','14':'Water','15':'Water','16':'Water','17':'Water',
    '18':'Waste','19':'Waste','20':'Waste',
    '21':'Health & Education',
}

VULN_SHEETS = [
    ('F_Vuln_Depth', 'Flood',      'Inundation depth (m)'),
    ('E_Vuln_PGA ',  'Earthquake', 'PGA (g)'),
    ('W_Vuln_V10m',  'Wind',       'Wind speed at 10 m (m/s)'),
    ('L_Vuln_PGD',   'Landslide',  'PGD (m)'),
]

FRAG_SHEETS = [
    ('E_Frag_PGA',   'Earthquake', 'PGA (g)'),
    ('L_Frag_PGD',   'Landslide',  'PGD (m)'),
]

DS_LABELS = ['Slight','Moderate','Extensive','Complete','Extensive/Complete']

def _ci_system(cid):
    try:
        num = re.sub(r'^[FEWL]','',cid).split('.')[0]
        return CI_SYSTEM.get(num,'Other')
    except: return 'Other'

def _base_id(cid):
    return cid[:-1] if re.search(r'\d+[ab]$',cid) else cid

def _is_bound(cid):
    if re.search(r'\d+a$',cid): return 'lo'
    if re.search(r'\d+b$',cid): return 'hi'
    return None

def _x_vals(data_rows):
    xs=[]
    for row in data_rows:
        if row[0] is None: continue
        try: xs.append(float(row[0]))
        except: continue
    return xs

def _y_col(data_rows, col, xs):
    ys=[]
    xi=0
    for row in data_rows:
        if row[0] is None: continue
        try: float(row[0])
        except: continue
        try: y=float(row[col]) if row[col] is not None else None
        except: y=None
        ys.append(y)
        xi+=1
    # align with xs, filter None pairs
    pairs=[(xs[i],ys[i]) for i in range(min(len(xs),len(ys))) if ys[i] is not None]
    return [round(p[0],5) for p in pairs],[round(p[1],5) for p in pairs]

def _parse_vuln_sheet(ws, hazard, im_label):
    rows=list(ws.iter_rows(values_only=True))
    if len(rows)<6: return []
    ids,descs,chars=rows[0],rows[1],rows[2]
    data_rows=rows[5:]
    xs_all=_x_vals(data_rows)

    all_cols={}
    for i in range(1,len(ids)):
        if ids[i] and str(ids[i]).strip():
            cid=str(ids[i]).strip()
            desc=str(descs[i]).strip() if descs[i] else ''
            char=str(chars[i]).strip() if chars[i] and str(chars[i]) not in ('None','N/A','nan') else ''
            all_cols[i]=(cid,desc,char)

    base_curves={}
    for col,(cid,desc,char) in all_cols.items():
        bid=_base_id(cid); role=_is_bound(cid)
        cx,cy=_y_col(data_rows,col,xs_all)
        if len(cx)<2: continue
        if role is None:
            if bid not in base_curves:
                base_curves[bid]={'id':bid,'hazard':hazard,'curve_type':'Vulnerability',
                    'im':im_label,'system':_ci_system(bid),'asset':desc,'details':char,
                    'label':desc+(' \u2014 '+char if char else ''),
                    'x':cx,'y':cy,'y_lo':None,'y_hi':None}
            else:
                base_curves[bid]['x']=cx; base_curves[bid]['y']=cy
        elif role=='lo':
            if bid not in base_curves:
                base_curves[bid]={'id':bid,'hazard':hazard,'curve_type':'Vulnerability',
                    'im':im_label,'system':_ci_system(bid),'asset':desc,'details':char,
                    'label':desc+(' \u2014 '+char if char else ''),
                    'x':cx,'y':None,'y_lo':cy,'y_hi':None}
            else: base_curves[bid]['y_lo']=cy
        else:
            if bid in base_curves: base_curves[bid]['y_hi']=cy

    result=[]
    for bid,rec in base_curves.items():
        if rec['y'] is None:
            if rec['y_lo'] is not None: rec['y']=rec['y_lo']
            else: continue
        result.append(rec)
    return result


def _parse_frag_sheet(ws, hazard, im_label):
    """Parse fragility sheet with multiple damage state columns per curve."""
    rows=list(ws.iter_rows(values_only=True))
    if len(rows)<6: return []
    ids,descs,chars,_,hdrs=rows[0],rows[1],rows[2],rows[3],rows[4]
    data_rows=rows[5:]
    xs_all=_x_vals(data_rows)

    # Group columns by base curve ID, collecting DS columns
    # Each unique base ID has columns for Slight/Moderate/Extensive/Complete
    curve_cols={}  # base_id -> {ds_label: col_index}
    cur_base=None; cur_desc=''; cur_char=''
    for i in range(1,len(ids)):
        if ids[i] and str(ids[i]).strip():
            cid=str(ids[i]).strip()
            bid=_base_id(cid)
            if _is_bound(cid): continue  # skip a/b variants in fragility sheets
            cur_base=bid
            cur_desc=str(descs[i]).strip() if descs[i] else ''
            cur_char=str(chars[i]).strip() if chars[i] and str(chars[i]) not in ('None','N/A','nan') else ''
            if bid not in curve_cols:
                curve_cols[bid]={'desc':cur_desc,'char':cur_char,'ds':{}}
        # Assign this column to a DS label
        if cur_base and hdrs[i]:
            ds=str(hdrs[i]).strip()
            if ds and ds not in ('Peak Ground Acceleration (g)','Peak Ground Deformation (cm)',
                                  'Water Intensity Measure (WIM)*','Wind speed at 10 m height (m/s)'):
                bid=cur_base
                if bid not in curve_cols:
                    curve_cols[bid]={'desc':cur_desc,'char':cur_char,'ds':{}}
                # Only take first occurrence of each DS label per curve
                if ds not in curve_cols[bid]['ds']:
                    curve_cols[bid]['ds'][ds]=i

    result=[]
    for bid,info in curve_cols.items():
        if not info['ds']: continue
        # Build DS arrays
        ds_data={}
        for ds_label,col in info['ds'].items():
            cx,cy=_y_col(data_rows,col,xs_all)
            if len(cx)>=2:
                ds_data[ds_label]={'x':cx,'y':cy}
        if not ds_data: continue

        # Use the x from the first DS
        first_ds=list(ds_data.values())[0]
        desc=info['desc']; char=info['char']
        result.append({
            'id':        bid,
            'hazard':    hazard,
            'curve_type':'Fragility',
            'im':        im_label,
            'system':    _ci_system(bid),
            'asset':     desc,
            'details':   char,
            'label':     desc+(' \u2014 '+char if char else ''),
            'x':         first_ds['x'],
            'y':         first_ds['y'],   # Slight = first DS (main curve for table)
            'y_lo':      None,
            'y_hi':      None,
            'ds':        {k:v['y'] for k,v in ds_data.items()},
        })
    return result


def extract_ci(base_dir):
    path=os.path.join(base_dir,'Table_D2_Hazard_Fragility_and_Vulnerability_Curves_V1.1.0.xlsx')
    print(f'  Reading CI vulnerability database from: {path}')
    if not os.path.exists(path):
        print(f'  ERROR: file not found: {path}'); return []

    wb=openpyxl.load_workbook(path,read_only=True,data_only=True)
    all_curves=[]

    for sname,hazard,im in VULN_SHEETS:
        ws=None
        for c in [sname,sname.strip(),sname.strip()+' ']:
            if c in wb.sheetnames: ws=wb[c]; break
        if not ws: print(f'    WARNING: {sname} not found'); continue
        curves=_parse_vuln_sheet(ws,hazard,im)
        all_curves.extend(curves)
        print(f'    {sname.strip():<20} Vulnerability  {len(curves):>4} curves')

    for sname,hazard,im in FRAG_SHEETS:
        ws=None
        for c in [sname,sname.strip()]:
            if c in wb.sheetnames: ws=wb[c]; break
        if not ws: print(f'    WARNING: {sname} not found'); continue
        curves=_parse_frag_sheet(ws,hazard,im)
        all_curves.extend(curves)
        ds_counts=[len(c.get('ds',{})) for c in curves]
        print(f'    {sname.strip():<20} Fragility      {len(curves):>4} curves (avg {sum(ds_counts)/max(len(ds_counts),1):.1f} DS)')

    wb.close()
    v=sum(1 for c in all_curves if c['curve_type']=='Vulnerability')
    f=sum(1 for c in all_curves if c['curve_type']=='Fragility')
    print(f'    Total: {len(all_curves)} curves ({v} vulnerability, {f} fragility)')
    return all_curves
